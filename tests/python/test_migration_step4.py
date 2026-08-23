from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
import pytest

from scripts.migrate_trip_data_step4 import (
    KOYASAN_NEW_NAME,
    KOYASAN_OLD_NAME,
    MigrationError,
    PLACEHOLDER_STAGE_ORDERS,
    TRANSPORT_DATE_CORRECTIONS,
    TRANSPORT_IDS,
    main,
    migrate_workbook,
)
from scripts.trip_data.validation import REQUIRED_SCHEMAS, validate_workbook


PRIVATE_BOOKING_VALUE = "https://private.example/path?private-token=do-not-print"

STAGES = [
    (1, "Tokyo", "東京", date(2026, 9, 11), date(2026, 9, 13)),
    (2, "Tokyo", "東京", date(2026, 9, 13), date(2026, 9, 14)),
    (3, "Tokyo", "東京", date(2026, 9, 14), date(2026, 9, 15)),
    (4, "Kamakura", "鎌倉", date(2026, 9, 15), date(2026, 9, 16)),
    (5, "Hakone", "箱根", date(2026, 9, 16), date(2026, 9, 17)),
    (6, "Kyoto", "京都", date(2026, 9, 17), date(2026, 9, 21)),
    (7, "Osaka", "大阪", date(2026, 9, 21), date(2026, 9, 23)),
    (8, "Koyasan", KOYASAN_OLD_NAME, date(2026, 9, 23), date(2026, 9, 25)),
    (9, "Wakayama", "和歌山", date(2026, 9, 25), date(2026, 9, 26)),
    (10, "Nara", "奈良", date(2026, 9, 26), date(2026, 9, 27)),
    (11, "Hiroshima", "広島", date(2026, 9, 27), date(2026, 9, 30)),
    (12, "Miyajima", "宮島", date(2026, 9, 30), date(2026, 10, 1)),
    (13, "Fukuoka", "福岡", date(2026, 10, 1), date(2026, 10, 3)),
    (14, "Tokyo", "東京", date(2026, 10, 3), date(2026, 10, 4)),
]


def _row(headers: list[str], **values: object) -> list[object | None]:
    return [values.get(header) for header in headers]


def create_migration_workbook(
    tmp_path: Path,
    *,
    include_transport_id: bool = False,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    stage_headers = list(REQUIRED_SCHEMAS["Stages"])
    stages_sheet = workbook.create_sheet("Stages")
    stages_sheet.append(stage_headers)
    for order, city, japanese_name, start_date, end_date in STAGES:
        stages_sheet.append(
            _row(
                stage_headers,
                Stage_Order=order,
                City=city,
                Japanese_Name=japanese_name,
                Start_Date=start_date,
                End_Date=end_date,
                Nights=(end_date - start_date).days,
            )
        )

    transport_headers = [
        column
        for column in REQUIRED_SCHEMAS["Transport"]
        if column != "Transport_ID"
    ] + ["Our Notes"]
    if include_transport_id:
        transport_headers.append("Transport_ID")
    transport_sheet = workbook.create_sheet("Transport")
    transport_sheet.append(transport_headers)
    stage_starts = {order: start for order, _, _, start, _ in STAGES}
    for stage_order in range(1, 15):
        values: dict[str, object] = {"Stage_Order": stage_order}
        if stage_order in TRANSPORT_IDS:
            values.update(
                {
                    "Date": TRANSPORT_DATE_CORRECTIONS.get(
                        stage_order,
                        (stage_starts[stage_order], stage_starts[stage_order]),
                    )[0],
                    "Mode": "Train",
                    "From": "Origin",
                    "To": "Destination",
                }
            )
        else:
            values["Our Notes"] = "Placeholder"
        transport_sheet.append(_row(transport_headers, **values))

    for sheet_name in ("Schedule", "Food", "Activities"):
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(list(REQUIRED_SCHEMAS[sheet_name]))

    hotel_headers = list(REQUIRED_SCHEMAS["Hotels"])
    hotels_sheet = workbook.create_sheet("Hotels")
    hotels_sheet.append(hotel_headers)
    hotels_sheet.append(
        _row(
            hotel_headers,
            Stage_Order=1,
            Hotel_Name="Reference Hotel",
            Booking_Link=PRIVATE_BOOKING_VALUE,
        )
    )
    list_sheet = workbook.create_sheet("List")
    list_sheet.append(["Type", "Status"])

    path = tmp_path / "migration.xlsx"
    workbook.save(path)
    workbook.close()
    return path


def transport_values(path: Path) -> tuple[list[str | None], dict[int, dict[str, object]]]:
    workbook = load_workbook(path, data_only=False)
    try:
        worksheet = workbook["Transport"]
        headers = [
            str(cell.value).strip() if cell.value is not None else None
            for cell in worksheet[1]
        ]
        rows: dict[int, dict[str, object]] = {}
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            record = {
                header: values[index]
                for index, header in enumerate(headers)
                if header is not None
            }
            stage_order = record["Stage_Order"]
            assert isinstance(stage_order, int)
            rows[stage_order] = record
        return headers, rows
    finally:
        workbook.close()


def workbook_values(path: Path) -> dict[tuple[str, str], object]:
    workbook = load_workbook(path, data_only=False)
    try:
        return {
            (worksheet.title, cell.coordinate): cell.value
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        }
    finally:
        workbook.close()


def set_formula_cache(path: Path, worksheet_xml: str, cell: str, value: int) -> None:
    replacement = path.with_name("formula-cache.xlsx")
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    cell_tag = f"{{{namespace}}}c"
    value_tag = f"{{{namespace}}}v"
    with ZipFile(path, "r") as source, ZipFile(
        replacement, "w", compression=ZIP_DEFLATED
    ) as destination:
        for item in source.infolist():
            content = source.read(item.filename)
            if item.filename == worksheet_xml:
                root = ElementTree.fromstring(content)
                target = next(
                    candidate
                    for candidate in root.iter(cell_tag)
                    if candidate.attrib.get("r") == cell
                )
                cached_value = target.find(value_tag)
                assert cached_value is not None
                cached_value.text = str(value)
                content = ElementTree.tostring(root, encoding="utf-8")
            destination.writestr(item, content)
    replacement.replace(path)


def test_missing_transport_id_column_is_added(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)

    summary = migrate_workbook(path)
    headers, _ = transport_values(path)

    assert summary.transport_id_column_added is True
    assert headers[-1] == "Transport_ID"


def test_existing_transport_id_column_is_reused(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path, include_transport_id=True)

    summary = migrate_workbook(path)
    headers, _ = transport_values(path)

    assert summary.transport_id_column_added is False
    assert headers.count("Transport_ID") == 1


def test_running_twice_does_not_duplicate_transport_id_column(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)

    migrate_workbook(path)
    migrate_workbook(path)
    headers, _ = transport_values(path)

    assert headers.count("Transport_ID") == 1


def test_exact_transport_id_mapping_is_applied(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)

    summary = migrate_workbook(path)
    _, rows = transport_values(path)

    assert summary.transport_ids_assigned == 10
    assert {
        stage_order: rows[stage_order]["Transport_ID"]
        for stage_order in TRANSPORT_IDS
    } == TRANSPORT_IDS


def test_stage_eleven_maps_to_tra008(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)

    migrate_workbook(path)
    _, rows = transport_values(path)

    assert rows[11]["Transport_ID"] == "TRA008"
    assert rows[11]["Transport_ID"] != "TRA011"


def test_transport_placeholders_remain_without_ids(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)

    migrate_workbook(path)
    _, rows = transport_values(path)

    assert all(
        rows[stage_order]["Transport_ID"] is None
        for stage_order in PLACEHOLDER_STAGE_ORDERS
    )


def test_five_known_transport_dates_are_corrected(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)

    summary = migrate_workbook(path)
    _, rows = transport_values(path)

    assert summary.transport_dates_corrected == 5
    assert all(
        rows[stage_order]["Date"].date() == corrected
        for stage_order, (_, corrected) in TRANSPORT_DATE_CORRECTIONS.items()
    )


def test_unexpected_transport_date_aborts_without_overwrite(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)
    workbook = load_workbook(path)
    worksheet = workbook["Transport"]
    worksheet.cell(2, 2).value = datetime(2026, 1, 1)
    workbook.save(path)
    workbook.close()

    with pytest.raises(MigrationError, match="Unexpected Transport date"):
        migrate_workbook(path)

    headers, rows = transport_values(path)
    assert "Transport_ID" not in headers
    assert rows[1]["Date"].date() == date(2026, 1, 1)


def test_koyasan_japanese_name_is_corrected(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)

    summary = migrate_workbook(path)
    values = workbook_values(path)

    assert summary.koyasan_name_corrected == 1
    assert values[("Stages", "C9")] == KOYASAN_NEW_NAME


def test_wakayama_japanese_name_remains_unchanged(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)

    migrate_workbook(path)
    values = workbook_values(path)

    assert values[("Stages", "C10")] == "和歌山"


def test_unexpected_stage_eight_identity_aborts(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)
    workbook = load_workbook(path)
    workbook["Stages"]["B9"] = "Unexpected City"
    workbook.save(path)
    workbook.close()

    with pytest.raises(MigrationError, match="not the expected Koyasan"):
        migrate_workbook(path)


def test_booking_link_value_remains_untouched(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)

    migrate_workbook(path)
    values = workbook_values(path)

    assert PRIVATE_BOOKING_VALUE in values.values()
    assert values[("Hotels", "F2")] == PRIVATE_BOOKING_VALUE


def test_migration_output_does_not_expose_booking_link(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    path = create_migration_workbook(tmp_path)

    assert main([str(path)]) == 0
    captured = capsys.readouterr()

    assert PRIVATE_BOOKING_VALUE not in captured.out
    assert PRIVATE_BOOKING_VALUE not in captured.err
    assert "Booking_Link" not in captured.out


def test_migrated_workbook_passes_existing_validator(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)

    migrate_workbook(path)
    report = validate_workbook(path)

    assert report.error_count == 0


def test_migration_is_semantically_idempotent(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)

    first_summary = migrate_workbook(path)
    first_values = workbook_values(path)
    second_summary = migrate_workbook(path)
    second_values = workbook_values(path)

    assert first_summary.transport_ids_assigned == 10
    assert second_summary.transport_ids_assigned == 0
    assert second_summary.transport_dates_corrected == 0
    assert second_summary.koyasan_name_corrected == 0
    assert second_values == first_values


def test_existing_formulas_are_preserved(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)
    workbook = load_workbook(path, data_only=False)
    workbook["Stages"]["F2"] = "=E2-D2"
    workbook.save(path)
    workbook.close()

    migrate_workbook(path)
    values = workbook_values(path)

    assert values[("Stages", "F2")] == "=E2-D2"


def test_cached_formula_values_are_preserved(tmp_path: Path) -> None:
    path = create_migration_workbook(tmp_path)
    workbook = load_workbook(path, data_only=False)
    workbook["Stages"]["F2"] = "=E2-D2"
    workbook.save(path)
    workbook.close()
    set_formula_cache(path, "xl/worksheets/sheet1.xml", "F2", 2)

    before = load_workbook(path, data_only=True)
    assert before["Stages"]["F2"].value == 2
    before.close()

    migrate_workbook(path)
    after = load_workbook(path, data_only=True)
    assert after["Stages"]["F2"].value == 2
    after.close()
