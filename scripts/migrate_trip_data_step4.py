"""Apply the audited Step 4 migration to the NihonNounours workbook."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date, datetime, time
import os
from pathlib import Path
import tempfile
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_WORKBOOK = Path("data/InputData_v2.xlsx")

TRANSPORT_IDS = {
    1: "TRA001",
    2: "TRA002",
    3: "TRA003",
    4: "TRA004",
    5: "TRA005",
    6: "TRA006",
    7: "TRA007",
    11: "TRA008",
    12: "TRA009",
    13: "TRA010",
}

PLACEHOLDER_STAGE_ORDERS = {8, 9, 10, 14}

TRANSPORT_DATE_CORRECTIONS = {
    1: (date(2026, 10, 11), date(2026, 9, 11)),
    2: (date(2026, 10, 13), date(2026, 9, 13)),
    3: (date(2026, 10, 14), date(2026, 9, 14)),
    4: (date(2026, 10, 15), date(2026, 9, 15)),
    5: (date(2026, 10, 16), date(2026, 9, 16)),
}

TRANSPORT_ENTITY_FIELDS = (
    "Date",
    "Start_Time",
    "End_Time",
    "Mode",
    "From",
    "To",
    "Service",
    "Status",
    "Info",
    "Important",
)

KOYASAN_OLD_NAME = "和歌山"
KOYASAN_NEW_NAME = "高野山"

SPREADSHEET_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


class MigrationError(RuntimeError):
    """Raised when workbook preconditions or safeguards are not satisfied."""


@dataclass(frozen=True)
class MigrationSummary:
    transport_id_column_added: bool
    transport_ids_assigned: int
    transport_dates_corrected: int
    koyasan_name_corrected: int


CellLocation = tuple[str, str]
WorkbookSnapshot = dict[CellLocation, object]


def _headers(worksheet: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in worksheet[1]:
        if cell.value is None:
            continue
        header = str(cell.value).strip()
        if not header:
            continue
        if header in headers:
            raise MigrationError(
                f"Duplicate header in {worksheet.title}: {header}"
            )
        headers[header] = cell.column
    return headers


def _required_column(headers: dict[str, int], sheet: str, column: str) -> int:
    try:
        return headers[column]
    except KeyError as exc:
        raise MigrationError(f"Missing required column: {sheet}.{column}") from exc


def _stage_rows(
    worksheet: Worksheet,
    stage_order_column: int,
) -> dict[int, int]:
    rows: dict[int, int] = {}
    for row_number in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row_number, stage_order_column).value
        if value is None:
            continue
        if not isinstance(value, int) or isinstance(value, bool):
            raise MigrationError(
                f"Unexpected Stage_Order type in {worksheet.title} row {row_number}"
            )
        if value in rows:
            raise MigrationError(
                f"Duplicate Stage_Order in {worksheet.title}: {value}"
            )
        rows[value] = row_number
    return rows


def _as_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _logical_snapshot(workbook: Workbook) -> WorkbookSnapshot:
    return {
        (worksheet.title, cell.coordinate): cell.value
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    }


def _snapshot_file(path: Path) -> WorkbookSnapshot:
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        return _logical_snapshot(workbook)
    finally:
        workbook.close()


def _assert_only_allowed_changes(
    before: WorkbookSnapshot,
    after: WorkbookSnapshot,
    allowed: set[CellLocation],
) -> None:
    changed = {
        location
        for location in before.keys() | after.keys()
        if before.get(location) != after.get(location)
    }
    unexpected = changed - allowed
    if unexpected:
        locations = ", ".join(
            f"{sheet}!{cell}" for sheet, cell in sorted(unexpected)
        )
        raise MigrationError(f"Unexpected logical workbook changes: {locations}")


def _worksheet_xml_names(archive: ZipFile) -> list[str]:
    return [
        name
        for name in archive.namelist()
        if name.startswith("xl/worksheets/") and name.endswith(".xml")
    ]


def _restore_formula_caches(source: Path, migrated: Path) -> None:
    """Restore untouched cached formula results discarded by openpyxl save."""

    formula_tag = f"{{{SPREADSHEET_NAMESPACE}}}f"
    value_tag = f"{{{SPREADSHEET_NAMESPACE}}}v"
    cell_tag = f"{{{SPREADSHEET_NAMESPACE}}}c"
    replacement_path = migrated.with_name(f".{migrated.name}.formulas.tmp")

    try:
        with ZipFile(source, "r") as source_zip, ZipFile(
            migrated, "r"
        ) as migrated_zip, ZipFile(
            replacement_path, "w", compression=ZIP_DEFLATED
        ) as output_zip:
            worksheet_names = set(_worksheet_xml_names(source_zip))
            for item in migrated_zip.infolist():
                content = migrated_zip.read(item.filename)
                if item.filename in worksheet_names:
                    source_root = ElementTree.fromstring(
                        source_zip.read(item.filename)
                    )
                    source_cells = {
                        cell.attrib.get("r"): cell
                        for cell in source_root.iter(cell_tag)
                        if cell.find(formula_tag) is not None
                    }
                    if not source_cells:
                        output_zip.writestr(item, content)
                        continue
                    migrated_root = ElementTree.fromstring(content)
                    for cell in migrated_root.iter(cell_tag):
                        coordinate = cell.attrib.get("r")
                        formula = cell.find(formula_tag)
                        if formula is None:
                            continue
                        source_cell = source_cells.get(coordinate)
                        if source_cell is None:
                            raise MigrationError(
                                "Formula cell was introduced unexpectedly"
                            )
                        source_formula = source_cell.find(formula_tag)
                        if source_formula is None:
                            raise MigrationError(
                                "An existing formula changed unexpectedly"
                            )
                        # Excel shared-formula followers omit their expression;
                        # openpyxl expands them to explicit equivalent formulas.
                        if (
                            source_formula.text is not None
                            and source_formula.text != formula.text
                        ):
                            raise MigrationError(
                                "An existing formula changed unexpectedly"
                            )
                        source_value = source_cell.find(value_tag)
                        if source_value is not None and source_value.text is not None:
                            migrated_value = cell.find(value_tag)
                            if migrated_value is None:
                                migrated_value = ElementTree.SubElement(
                                    cell, value_tag
                                )
                            migrated_value.text = source_value.text
                    content = ElementTree.tostring(
                        migrated_root,
                        encoding="utf-8",
                        xml_declaration=False,
                    )
                output_zip.writestr(item, content)
        replacement_path.replace(migrated)
    finally:
        replacement_path.unlink(missing_ok=True)


def _apply_transport_migration(
    worksheet: Worksheet,
) -> tuple[bool, int, int, set[CellLocation]]:
    headers = _headers(worksheet)
    stage_order_column = _required_column(
        headers, worksheet.title, "Stage_Order"
    )
    date_column = _required_column(headers, worksheet.title, "Date")
    for column in TRANSPORT_ENTITY_FIELDS:
        _required_column(headers, worksheet.title, column)

    column_added = "Transport_ID" not in headers
    if column_added:
        transport_id_column = worksheet.max_column + 1
        worksheet.cell(1, transport_id_column).value = "Transport_ID"
    else:
        transport_id_column = headers["Transport_ID"]

    rows = _stage_rows(worksheet, stage_order_column)
    expected_orders = set(TRANSPORT_IDS) | PLACEHOLDER_STAGE_ORDERS
    if set(rows) != expected_orders:
        raise MigrationError(
            "Transport Stage_Order rows differ from the approved Step 4 structure"
        )

    entity_columns = [headers[column] for column in TRANSPORT_ENTITY_FIELDS]
    for stage_order in TRANSPORT_IDS:
        row_number = rows[stage_order]
        if not any(
            worksheet.cell(row_number, column).value not in (None, "")
            for column in entity_columns
        ):
            raise MigrationError(
                f"Expected meaningful Transport row for Stage {stage_order}"
            )
    for stage_order in PLACEHOLDER_STAGE_ORDERS:
        row_number = rows[stage_order]
        if any(
            worksheet.cell(row_number, column).value not in (None, "")
            for column in entity_columns
        ):
            raise MigrationError(
                f"Expected Transport placeholder row for Stage {stage_order}"
            )

    allowed: set[CellLocation] = {
        (worksheet.title, worksheet.cell(1, transport_id_column).coordinate)
    }
    assigned = 0
    for stage_order, transport_id in TRANSPORT_IDS.items():
        cell = worksheet.cell(rows[stage_order], transport_id_column)
        if cell.value is None:
            cell.value = transport_id
            assigned += 1
        elif cell.value != transport_id:
            raise MigrationError(
                f"Unexpected Transport_ID for Stage {stage_order}"
            )
        allowed.add((worksheet.title, cell.coordinate))

    for stage_order in PLACEHOLDER_STAGE_ORDERS:
        cell = worksheet.cell(rows[stage_order], transport_id_column)
        if cell.value not in (None, ""):
            raise MigrationError(
                f"Transport placeholder Stage {stage_order} has an unexpected ID"
            )
        allowed.add((worksheet.title, cell.coordinate))

    corrected_dates = 0
    for stage_order, (expected_old, corrected) in TRANSPORT_DATE_CORRECTIONS.items():
        cell = worksheet.cell(rows[stage_order], date_column)
        current_date = _as_date(cell.value)
        if current_date == expected_old:
            cell.value = datetime.combine(corrected, time.min)
            corrected_dates += 1
        elif current_date != corrected:
            raise MigrationError(
                f"Unexpected Transport date for Stage {stage_order}"
            )
        allowed.add((worksheet.title, cell.coordinate))

    return column_added, assigned, corrected_dates, allowed


def _apply_koyasan_migration(
    worksheet: Worksheet,
) -> tuple[int, set[CellLocation]]:
    headers = _headers(worksheet)
    stage_order_column = _required_column(
        headers, worksheet.title, "Stage_Order"
    )
    city_column = _required_column(headers, worksheet.title, "City")
    japanese_name_column = _required_column(
        headers, worksheet.title, "Japanese_Name"
    )
    rows = _stage_rows(worksheet, stage_order_column)
    if 8 not in rows:
        raise MigrationError("Stages Stage_Order 8 is missing")

    row_number = rows[8]
    if worksheet.cell(row_number, city_column).value != "Koyasan":
        raise MigrationError("Stage 8 is not the expected Koyasan row")

    cell = worksheet.cell(row_number, japanese_name_column)
    corrected = 0
    if cell.value == KOYASAN_OLD_NAME:
        cell.value = KOYASAN_NEW_NAME
        corrected = 1
    elif cell.value != KOYASAN_NEW_NAME:
        raise MigrationError("Stage 8 has an unexpected Japanese_Name")

    return corrected, {(worksheet.title, cell.coordinate)}


def migrate_workbook(path: str | Path = DEFAULT_WORKBOOK) -> MigrationSummary:
    """Apply the controlled migration atomically and return change counts."""

    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise MigrationError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    temporary_path: Path | None = None
    try:
        if "Transport" not in workbook.sheetnames or "Stages" not in workbook.sheetnames:
            raise MigrationError("Required migration sheets are missing")

        before = _logical_snapshot(workbook)
        column_added, assigned, corrected_dates, transport_allowed = (
            _apply_transport_migration(workbook["Transport"])
        )
        corrected_name, stage_allowed = _apply_koyasan_migration(
            workbook["Stages"]
        )
        after = _logical_snapshot(workbook)
        _assert_only_allowed_changes(
            before,
            after,
            transport_allowed | stage_allowed,
        )

        descriptor, temporary_name = tempfile.mkstemp(
            dir=workbook_path.parent,
            prefix=f".{workbook_path.stem}.step4.",
            suffix=".xlsx",
        )
        os.close(descriptor)
        temporary_path = Path(temporary_name)
        workbook.save(temporary_path)
        workbook.close()

        _restore_formula_caches(workbook_path, temporary_path)
        saved = _snapshot_file(temporary_path)
        if saved != after:
            raise MigrationError("Saved workbook logical values changed unexpectedly")

        temporary_path.replace(workbook_path)
        temporary_path = None
        return MigrationSummary(
            transport_id_column_added=column_added,
            transport_ids_assigned=assigned,
            transport_dates_corrected=corrected_dates,
            koyasan_name_corrected=corrected_name,
        )
    finally:
        workbook.close()
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Apply the controlled NihonNounours Step 4 workbook migration."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"Workbook path (default: {DEFAULT_WORKBOOK.as_posix()})",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    print("NihonNounours Step 4 migration")
    print()
    try:
        summary = migrate_workbook(args.workbook)
    except Exception as exc:
        print(f"Migration aborted: {exc}")
        return 1

    column_status = (
        "added" if summary.transport_id_column_added else "reused"
    )
    print(f"Transport_ID column: {column_status}")
    print(f"Transport IDs assigned: {summary.transport_ids_assigned}")
    print(f"Transport dates corrected: {summary.transport_dates_corrected}")
    print(f"Koyasan Japanese name corrected: {summary.koyasan_name_corrected}")
    print()
    print("Migration completed successfully.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
