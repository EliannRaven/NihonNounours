"""Read the trip workbook safely and normalize low-level cell values."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Final

from openpyxl import load_workbook


DATE_COLUMNS: Final = {
    ("Stages", "Start_Date"),
    ("Stages", "End_Date"),
    ("Transport", "Date"),
    ("Schedule", "Date"),
}

TIME_COLUMNS: Final = {
    ("Hotels", "Checkin_Time"),
    ("Hotels", "Checkout_Time"),
    ("Transport", "Start_Time"),
    ("Transport", "End_Time"),
    ("Schedule", "Start_Time"),
    ("Schedule", "End_Time"),
}


class WorkbookReadError(RuntimeError):
    """Raised when a workbook cannot be opened or read safely."""


@dataclass(frozen=True)
class WorkbookRow:
    """One non-blank worksheet row with its original Excel row number."""

    number: int
    values: dict[str, object | None]

    def get(self, column: str) -> object | None:
        return self.values.get(column)


@dataclass(frozen=True)
class SheetData:
    """Normalized headers and meaningful rows from one worksheet."""

    name: str
    headers: tuple[str, ...]
    duplicate_headers: frozenset[str]
    rows: tuple[WorkbookRow, ...]

    def has_column(self, column: str) -> bool:
        return column in self.headers


@dataclass(frozen=True)
class WorkbookData:
    """In-memory, read-only representation of the workbook data."""

    path: Path
    sheets: dict[str, SheetData]

    def get_sheet(self, name: str) -> SheetData | None:
        return self.sheets.get(name)


def normalize_header(sheet: str, value: object | None) -> str | None:
    """Trim a header and apply only explicitly supported header aliases."""

    if value is None:
        return None
    header = str(value).strip()
    if not header:
        return None
    if sheet == "Transport" and header == "Our_Notes":
        return "Our Notes"
    return header


def normalize_value(
    value: object | None,
    *,
    sheet: str,
    column: str | None,
) -> object | None:
    """Conservatively normalize a cell without guessing malformed values."""

    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None

    if column is not None and (sheet, column) in DATE_COLUMNS:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

    if column is not None and (sheet, column) in TIME_COLUMNS:
        if isinstance(value, datetime):
            return value.time().replace(tzinfo=None)
        if isinstance(value, time):
            return value.replace(tzinfo=None)

    if isinstance(value, float) and value.is_integer():
        return int(value)

    return value


def read_workbook(path: str | Path) -> WorkbookData:
    """Open an xlsx workbook read-only and return normalized in-memory data."""

    workbook_path = Path(path)
    workbook = None
    try:
        workbook = load_workbook(
            workbook_path,
            read_only=True,
            data_only=True,
        )
        sheets: dict[str, SheetData] = {}

        for worksheet in workbook.worksheets:
            raw_rows = worksheet.iter_rows(values_only=True)
            first_row = next(raw_rows, ())
            normalized_headers = [
                normalize_header(worksheet.title, value) for value in first_row
            ]

            seen_headers: set[str] = set()
            duplicates: set[str] = set()
            for header in normalized_headers:
                if header is None:
                    continue
                if header in seen_headers:
                    duplicates.add(header)
                seen_headers.add(header)

            rows: list[WorkbookRow] = []
            for row_number, raw_values in enumerate(raw_rows, start=2):
                values: dict[str, object | None] = {}
                is_blank = True

                for index, raw_value in enumerate(raw_values):
                    header = (
                        normalized_headers[index]
                        if index < len(normalized_headers)
                        else None
                    )
                    value = normalize_value(
                        raw_value,
                        sheet=worksheet.title,
                        column=header,
                    )
                    if value is not None:
                        is_blank = False
                    if header is not None and header not in values:
                        values[header] = value

                if not is_blank:
                    rows.append(WorkbookRow(number=row_number, values=values))

            sheets[worksheet.title] = SheetData(
                name=worksheet.title,
                headers=tuple(header for header in normalized_headers if header),
                duplicate_headers=frozenset(duplicates),
                rows=tuple(rows),
            )

        return WorkbookData(path=workbook_path, sheets=sheets)
    except Exception as exc:
        raise WorkbookReadError(
            f"Could not read workbook: {workbook_path}"
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()
